from openpyxl import Workbook, load_workbook
import os

class UserDatabase:
    def __init__(self, file_path="users.xlsx"):
        self.file_path = file_path
        self.sheet_name = "Users"
        self._initialize_excel()

    def _initialize_excel(self):
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["Username", "Password"])
            wb.save(self.file_path)



    def _load_workbook(self):
        return load_workbook(self.file_path)

    def add_user(self, username, password):
        if self.user_exists(username):
            return False  # user already exists

        wb = self._load_workbook()
        ws = wb[self.sheet_name]
        ws.append([username, password])
        wb.save(self.file_path)
        return True

    def user_exists(self, username):
        wb = self._load_workbook()
        ws = wb[self.sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                return True
        return False

    def validate_user(self, username, password):
        wb = self._load_workbook()
        ws = wb[self.sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                return True
        return False
